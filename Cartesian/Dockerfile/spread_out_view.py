import nibabel
import numpy as np
from matplotlib import colors, pyplot as plt
import os
import matplotlib
import re
import pandas as pd
import openpyxl
from ct2polar import lumen_center, sample_line, holes_detect, fill_holes
import math
from matplotlib.ticker import MultipleLocator
matplotlib.use("Agg")

def save_nii_gz(savepath,img):
    pair_img = nibabel.Nifti1Pair(img,np.eye(4))
    nibabel.save(pair_img,savepath)

def chemogram_plot_RGB_expert_IVUS(chemogram_save_path,df=None,test_vessels_name=None,angle_interval=10):
    '''plot chemogram'''
    # test_vessels_name = ['01_LAD_DL','01_LCX_DL','02_LAD_DL','02_LCX_DL','02_RCA_DL','03_LAD_DL','03_LCX_DL','03_RCA_DL']

    for vessel in test_vessels_name:

        pred_lipid = nibabel.load(os.path.join(chemogram_save_path,vessel+'_lipid.nii.gz')).get_fdata()
        pred_calcium = nibabel.load(os.path.join(chemogram_save_path,vessel+'_calcium.nii.gz')).get_fdata()

        # test_slice_idx = []
        # row_idx = []
        # cnt = 0
        # angle_interval = 10 #deg
        # # for slice_idx in df['CTSliceIdx']:
        # for slice_idx in df['Ctsliceidx']:
        #     # if df['VesselName'][cnt]==vessel+'_DL':
        #     if df['Vesselname'][cnt]==vessel+'_DL.zip':
        #         if slice_idx>=label_lipid.shape[1]:
        #             break
        #         test_slice_idx.append(int(slice_idx))
        #         row_idx.append(cnt)
        #     cnt = cnt+1
        # shadow = np.zeros((int(label_lipid.shape[0]/angle_interval),label_lipid.shape[1]))
        # shadow[:,test_slice_idx] = 2
        # shadow = 2-shadow
        
        pred_lipid_squeeze = np.zeros((int(pred_lipid.shape[0]/angle_interval),pred_lipid.shape[1]))   
        pred_calcium_squeeze = np.zeros((int(pred_calcium.shape[0]/angle_interval),pred_calcium.shape[1])) 

        # row_cnt = 0
        for row in range(pred_lipid.shape[0]):
            # row_cnt = row_cnt+1
            tmp_pred_lipid = pred_lipid.copy()
            tmp_pred_lipid[pred_lipid==-0.1] = 0
            tmp_pred_calcium = pred_calcium.copy()
            tmp_pred_calcium[pred_calcium==-0.1] = 0
            if (row+1)%angle_interval==0:
                pred_lipid_squeeze[int(row/angle_interval),:] = np.mean(tmp_pred_lipid[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                pred_calcium_squeeze[int(row/angle_interval),:] = np.mean(tmp_pred_calcium[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
        pred_lipid_squeeze = np.round(pred_lipid_squeeze,decimals=1)
        pred_calcium_squeeze = np.round(pred_calcium_squeeze,decimals=1)
        
        # start_slice = test_slice_idx[0]
        # end_slice = test_slice_idx[-1]

        os.makedirs(os.path.join(chemogram_save_path.replace('_Chemogram','_Chemogram_png'), vessel), exist_ok=True)
        '''squeezed chemogram (angle interval )'''
        plt.rcParams.update({'font.size': 12})
        fig_, axs_ = plt.subplots(1,1)
        ax1_ = plt.subplot(1,1,1)
        plt.contourf(pred_lipid_squeeze,levels=[0,0.5,1],colors=('r','y'))
        # ax1_.set_xlim((start_slice,end_slice))
        ax1_.set_title('lipid')
        ax1_.set_ylabel('ground truth')
        plt.savefig(os.path.join(chemogram_save_path.replace('_Chemogram','_Chemogram_png'), vessel, vessel+'_lipid_pred_'+str(angle_interval)+'.png'))

        plt.rcParams.update({'font.size': 12})
        fig_, axs_ = plt.subplots(1,1)
        ax1_ = plt.subplot(1,1,1)
        plt.contourf(pred_calcium_squeeze,levels=[0,0.5,1],colors=('r','y'))
        # ax1_.set_xlim((start_slice,end_slice))
        ax1_.set_title('calcium')
        ax1_.set_ylabel('ground truth')
        plt.savefig(os.path.join(chemogram_save_path.replace('_Chemogram','_Chemogram_png'), vessel, vessel+'_calcium_pred_'+str(angle_interval)+'.png'))

        # plt.show()

# Comparison
def chemogram_plot_final(chemogram_save_path,expert_chemogram_save_path=None,df=None,chemo_png_save_path=None): # showing prediction of 2.5D Dense U-Net and 2.5D Mask RCNN at the same time 
    matplotlib.use('Agg')
    # if not os.path.exists(chemo_png_save_path):
    #     os.makedirs(chemo_png_save_path)
    '''plot chemogram'''
    # test_vessels_name = ['01_LAD_DL','01_LCX_DL','02_LAD_DL','02_LCX_DL','02_RCA_DL','03_LAD_DL','03_LCX_DL','03_RCA_DL']
    test_vessels_name = ['02_LAD_final_DL', '04_LAD_DLupdated']

    levels = np.arange(0,1.2,0.2)
    levels2 = np.arange(0,1.2,0.2)

    for vessel in test_vessels_name[:]:
        '## load predicted chemogram of 2.5D Dense U-Net'
        pred_lipid_1 = nibabel.load(os.path.join(chemogram_save_path[0],vessel+'_lipid.nii.gz')).get_fdata()
        pred_calcium_1 = nibabel.load(os.path.join(chemogram_save_path[0],vessel+'_calcium.nii.gz')).get_fdata()
        '## load predicted chemogram of 2.5D Mask RCNN'
        pred_lipid_2 = nibabel.load(os.path.join(chemogram_save_path[1],vessel+'_lipid.nii.gz')).get_fdata()
        pred_calcium_2 = nibabel.load(os.path.join(chemogram_save_path[1],vessel+'_calcium.nii.gz')).get_fdata()
        # label_lipid = nibabel.load(os.path.join(chemogram_save_path[0],vessel+'_lipid_label.nii.gz')).get_fdata()[:,:pred_lipid_1.shape[1]]
        # label_calcium = nibabel.load(os.path.join(chemogram_save_path[0],vessel+'_calcium_label.nii.gz')).get_fdata()[:,:pred_calcium_1.shape[1]]
        # expert_lipid = nibabel.load(os.path.join(expert_chemogram_save_path,vessel.split('_')[0]+'_'+vessel.split('_')[1]+'_lipid_label.nii.gz')).get_fdata()[:,:pred_lipid_1.shape[1]]
        # expert_calcium = nibabel.load(os.path.join(expert_chemogram_save_path,vessel.split('_')[0]+'_'+vessel.split('_')[1]+'_calcium_label.nii.gz')).get_fdata()[:,:pred_calcium_1.shape[1]]

        ####  generate Siemens metrics only ####
        # test_slice_idx_1 = []
        # row_idx = []
        # cnt = 0
        angle_interval = 30
        # for slice_idx in df['Ctsliceidx']:
        #     if df['Vesselname'][cnt].split('.')[0]==vessel:
        #         if slice_idx>=pred_lipid_1.shape[1]:
        #             break
        #         test_slice_idx_1.append(int(slice_idx))
        #         row_idx.append(cnt)
        #     cnt = cnt+1
        # shadow_1 = np.zeros((int(pred_lipid_1.shape[0]/angle_interval),pred_lipid_1.shape[1]))
        # shadow_1[:,test_slice_idx_1] = 2
        # shadow_1 = 2-shadow_1

        # test_slice_idx_2 = []
        # row_idx = []
        # cnt = 0
        # angle_interval = 30
        # for slice_idx in df['Ctsliceidx']:
        #     if df['Vesselname'][cnt].split('.')[0]==vessel:
        #         if slice_idx>=pred_lipid_2.shape[1]:
        #             break
        #         test_slice_idx_2.append(int(slice_idx))
        #         row_idx.append(cnt)
        #     cnt = cnt+1
        # shadow_2 = np.zeros((int(pred_lipid_2.shape[0]/angle_interval),pred_lipid_2.shape[1]))
        # shadow_2[:,test_slice_idx_2] = 2
        # shadow_2 = 2-shadow_2

        ####  generate GE and Siemens metrics at the same time ####
        # test_slice_idx = []
        # row_idx = []
        # cnt = 0
        # angle_interval = 30
        # if (vessel.split('_')[0]+'_'+vessel.split('_')[1]) in list(set(df[1]['Vesselname'])): # for GE data
        #     for slice_idx in df[1]['Ctsliceidx']:
        #         if str(df[1]['Vesselname'][cnt])=='nan':
        #             break
        #         if df[1]['Vesselname'][cnt] in vessel:
        #             if slice_idx>=pred_lipid.shape[1]:
        #                 break
        #             test_slice_idx.append(int(slice_idx))
        #             row_idx.append(cnt)
        #         cnt = cnt+1
        # elif vessel+'.zip' in list(set(df[0]['Vesselname'])): # for Siemens data
        #     for slice_idx in df[0]['Ctsliceidx']:
        #         if df[0]['Vesselname'][cnt].split('.')[0] in vessel:
        #             if slice_idx>=pred_lipid.shape[1]:
        #                 break
        #             test_slice_idx.append(int(slice_idx))
        #             row_idx.append(cnt)
        #         cnt = cnt+1

        # shadow = np.zeros((int(pred_lipid.shape[0]/angle_interval),pred_lipid.shape[1]))
        # shadow[:,test_slice_idx] = 2
        # shadow = 2-shadow
        
        show_thre_lipid = 0.6
        show_thre_calcium = 0.6

        pred_lipid_1[pred_lipid_1<show_thre_lipid+0.1] = 0
        pred_calcium_1[pred_calcium_1<show_thre_calcium+0.1] = 0
        # pred_lipid_2[pred_lipid_2<show_thre_lipid+0.1] = 0
        # pred_calcium_2[pred_calcium_2<show_thre_calcium+0.1] = 0

        pred_lipid_squeeze_1 = np.zeros((int(pred_lipid_1.shape[0]/angle_interval),pred_lipid_1.shape[1]))   
        pred_calcium_squeeze_1 = np.zeros((int(pred_calcium_1.shape[0]/angle_interval),pred_calcium_1.shape[1]))    
        pred_lipid_squeeze_2 = np.zeros((int(pred_lipid_2.shape[0]/angle_interval),pred_lipid_2.shape[1]))   
        pred_calcium_squeeze_2 = np.zeros((int(pred_calcium_2.shape[0]/angle_interval),pred_calcium_2.shape[1]))    
        # label_lipid_squeeze = np.zeros((int(label_lipid.shape[0]/angle_interval),label_lipid.shape[1]))   
        # label_calcium_squeeze = np.zeros((int(label_calcium.shape[0]/angle_interval),label_calcium.shape[1])) 
        # expert_lipid_squeeze = np.zeros((int(expert_lipid.shape[0]/angle_interval),expert_lipid.shape[1])) 
        # expert_calcium_squeeze = np.zeros((int(expert_calcium.shape[0]/angle_interval),expert_calcium.shape[1])) 
        # row_cnt = 0
        for row in range(pred_lipid_1.shape[0]):
            # row_cnt = row_cnt+1
            tmp_pred_lipid_1 = pred_lipid_1.copy()
            tmp_pred_lipid_1[pred_lipid_1==-0.1] = 0
            tmp_pred_calcium_1 = pred_calcium_1.copy()
            tmp_pred_calcium_1[pred_calcium_1==-0.1] = 0
            tmp_pred_lipid_2 = pred_lipid_2.copy()
            tmp_pred_lipid_2[pred_lipid_2==-0.1] = 0
            tmp_pred_calcium_2 = pred_calcium_2.copy()
            tmp_pred_calcium_2[pred_calcium_2==-0.1] = 0

            # tmp_label_lipid = label_lipid.copy()
            # tmp_label_lipid[label_lipid==-0.1] = 0
            # tmp_label_calcium = label_calcium.copy()
            # tmp_label_calcium[label_calcium==-0.1] = 0
            # tmp_expert_lipid = expert_lipid.copy()
            # tmp_expert_lipid[expert_lipid==-0.1] = 0
            # tmp_expert_calcium = expert_calcium.copy()
            # tmp_expert_calcium[expert_calcium==-0.1] = 0
            if (row+1)%angle_interval==0:
                pred_lipid_squeeze_1[int(row/angle_interval),:] = np.mean(tmp_pred_lipid_1[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                pred_calcium_squeeze_1[int(row/angle_interval),:] = np.mean(tmp_pred_calcium_1[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                pred_lipid_squeeze_2[int(row/angle_interval),:] = np.mean(tmp_pred_lipid_2[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                pred_calcium_squeeze_2[int(row/angle_interval),:] = np.mean(tmp_pred_calcium_2[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)    
                
                # label_lipid_squeeze[int(row/angle_interval),:] = np.mean(tmp_label_lipid[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                # label_calcium_squeeze[int(row/angle_interval),:] = np.mean(tmp_label_calcium[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                # expert_lipid_squeeze[int(row/angle_interval),:] = np.mean(tmp_expert_lipid[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                # expert_calcium_squeeze[int(row/angle_interval),:] = np.mean(tmp_expert_calcium[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)    
        pred_lipid_squeeze_1 = np.round(pred_lipid_squeeze_1,decimals=1)#+shadow_1
        pred_calcium_squeeze_1 = np.round(pred_calcium_squeeze_1,decimals=1)#+shadow_1
        pred_lipid_squeeze_2 = np.round(pred_lipid_squeeze_2,decimals=1)#+shadow_2
        pred_calcium_squeeze_2 = np.round(pred_calcium_squeeze_2,decimals=1)#+shadow_2

        # label_lipid_squeeze = np.round(label_lipid_squeeze,decimals=1)+shadow_1
        # label_calcium_squeeze = np.round(label_calcium_squeeze,decimals=1)+shadow_1
        # expert_lipid_squeeze = np.round(expert_lipid_squeeze,decimals=1)+shadow_1
        # expert_calcium_squeeze = np.round(expert_calcium_squeeze,decimals=1)+shadow_1

        # start_slice_1 = test_slice_idx_1[0]
        # end_slice_1 = test_slice_idx_1[-1]
        # start_slice_2 = test_slice_idx_2[0]
        # end_slice_2 = test_slice_idx_2[-1]

        
        '''unsqueezed chemogram (angle interval = 1deg)'''
        # # fig, axs = plt.subplots(2,2)
        # fig, axs = plt.subplots(4,1)
        # # ax1 = plt.subplot(2,2,1)
        # ax1 = plt.subplot(4,1,1)
        # h_pred_lipid = plt.contourf(pred_lipid, cmap=plt.cm.coolwarm, levels=levels)
        # c_pred_lipid = plt.colorbar(h_pred_lipid)
        # # ax1.set_title('lipid')
        # ax1.set_ylabel('lipid pred')

        # ax2 = plt.subplot(4,1,3)
        # h_pred_calcium = plt.contourf(pred_calcium, cmap=plt.cm.coolwarm, levels=levels)
        # c_pred_calcium = plt.colorbar(h_pred_calcium)
        # # ax2.set_title('calcium')
        # ax2.set_ylabel('Ca pred')

        # ax3 = plt.subplot(4,1,2)
        # h_label_lipid = plt.contourf(label_lipid, cmap=plt.cm.coolwarm, levels=levels)
        # c_label_lipid = plt.colorbar(h_label_lipid)
        # ax3.set_ylabel('lipid label')

        # ax4 = plt.subplot(4,1,4)
        # h_label_calcium = plt.contourf(label_calcium, cmap=plt.cm.coolwarm, levels=levels)
        # c_label_calcium = plt.colorbar(h_label_calcium)
        # ax4.set_ylabel('Ca label')
        # # print(1.0)

        # transversal_lipid = 0.5 #[0.4,0.6]
        # transversal_calcium = 0.9 #[0.8,1.0]

        # ''' calculate lipid metrix'''
        # y_pred_lipid = np.zeros((label_lipid.shape[0],label_lipid.shape[1]))
        # y_true_lipid = np.zeros((label_lipid.shape[0],label_lipid.shape[1]))
        # tmp1 = np.argwhere(pred_lipid>=transversal_lipid)
        # tmp2 = np.argwhere(label_lipid>=transversal_lipid)
        # y_pred_lipid[tmp1[:,0],tmp1[:,1]] = 1
        # y_true_lipid[tmp2[:,0],tmp2[:,1]] = 1
        # recall_high_prob1 = recall_m(y_true_lipid,y_pred_lipid)
        # precision_high_prob1 = precision_m(y_true_lipid,y_pred_lipid)

        # '''calculate calcium metrix'''
        # y_pred_calcium = np.zeros((label_calcium.shape[0],label_calcium.shape[1]))
        # y_true_calcium = np.zeros((label_calcium.shape[0],label_calcium.shape[1]))
        # tmp1 = np.argwhere(pred_calcium>=transversal_calcium)
        # tmp2 = np.argwhere(label_calcium>=transversal_calcium)
        # y_pred_calcium[tmp1[:,0],tmp1[:,1]] = 1
        # y_true_calcium[tmp2[:,0],tmp2[:,1]] = 1
        # recall_high_prob2 = recall_m(y_true_calcium,y_pred_calcium)
        # precision_high_prob2 = precision_m(y_true_calcium,y_pred_calcium)

        # try:    
        #     fig.suptitle(vessel+'\n'+'[lipid: precision '+str(format(precision_high_prob1,'.4f'))+' recall '+str(format(recall_high_prob1,'.4f'))+']'+'\n'
        #                         +'[calcium: precision '+str(format(precision_high_prob2,'.4f'))+' recall '+str(format(recall_high_prob2,'.4f'))+']')
        # except:
        #     fig.suptitle(vessel[0]+vessel[1]+'\n'+'[lipid: precision '+str(format(precision_high_prob1,'.4f'))+' recall '+str(format(recall_high_prob1,'.4f'))+']'+'\n'
        #                         +'[calcium: precision '+str(format(precision_high_prob2,'.4f'))+' recall '+str(format(recall_high_prob2,'.4f'))+']')   

        '''squeezed chemogram (angle interval = 30deg)'''
        plt.rcParams['figure.figsize'] = [10, 8]
        plt.rcParams.update({'font.size': 10})
        fig_, axs_ = plt.subplots(2,2)

        ax1_ = plt.subplot(2,2,1)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        plt.xticks(fontsize=7)
        plt.yticks(fontsize=7)
        h_pred_lipid_ = plt.contourf(pred_lipid_squeeze_1, cmap=plt.cm.coolwarm, levels=levels)
        c_pred_lipid_ = plt.colorbar(h_pred_lipid_)
        c_pred_lipid_.ax.tick_params(labelsize=7)
        # ax1_.set_xlim((start_slice_1,end_slice_1))
        ax1_.yaxis.set_major_locator(MultipleLocator(3))
        ax1_.set_yticklabels(['0','0','90','180','270'])
        ax1_.set_title('lipid',fontsize=8)
        ax1_.set_ylabel('Dense UNet',fontsize=8)

        ax2_ = plt.subplot(2,2,2)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        plt.xticks(fontsize=7)
        plt.yticks(fontsize=7)
        h_pred_calcium_ = plt.contourf(pred_calcium_squeeze_1, cmap=plt.cm.coolwarm, levels=levels)
        c_pred_calcium_ = plt.colorbar(h_pred_calcium_)
        c_pred_calcium_.ax.tick_params(labelsize=7)
        # ax2_.set_xlim((start_slice_1,end_slice_1))
        ax2_.yaxis.set_major_locator(MultipleLocator(3))
        ax2_.set_yticklabels(['0','0','90','180','270'])
        ax2_.set_title('calcium',fontsize=8)

        ax3_ = plt.subplot(2,2,3)
        # plt.contourf(shadow_2,cmap='Greens',levels=levels2)
        plt.xticks(fontsize=7)
        plt.yticks(fontsize=7)
        h_pred_lipid_ = plt.contourf(pred_lipid_squeeze_2, cmap=plt.cm.coolwarm, levels=levels)
        c_pred_lipid_ = plt.colorbar(h_pred_lipid_)
        c_pred_lipid_.ax.tick_params(labelsize=7)
        # ax3_.set_xlim((start_slice_2,end_slice_2))
        ax3_.yaxis.set_major_locator(MultipleLocator(3))
        ax3_.set_yticklabels(['0','0','90','180','270'])
        ax3_.set_ylabel('Mask RCNN',fontsize=8)

        ax4_ = plt.subplot(2,2,4)
        # plt.contourf(shadow_2,cmap='Greens',levels=levels2)
        plt.xticks(fontsize=7)
        plt.yticks(fontsize=7)
        h_pred_calcium_ = plt.contourf(pred_calcium_squeeze_2, cmap=plt.cm.coolwarm, levels=levels)
        c_pred_calcium_ = plt.colorbar(h_pred_calcium_)
        c_pred_calcium_.ax.tick_params(labelsize=7)
        # ax4_.set_xlim((start_slice_2,end_slice_2))
        ax4_.yaxis.set_major_locator(MultipleLocator(3))
        ax4_.set_yticklabels(['0','0','90','180','270'])

        # ax5_ = plt.subplot(4,2,5)
        # # plt.contourf(label_lipid_squeeze)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        # plt.xticks(fontsize=7)
        # plt.yticks(fontsize=7)
        # h_label_lipid_ = plt.contourf(label_lipid_squeeze, cmap=plt.cm.coolwarm, levels=levels)
        # c_label_lipid_ = plt.colorbar(h_label_lipid_)
        # c_label_lipid_.ax.tick_params(labelsize=7)
        # ax5_.set_xlim((start_slice_1,end_slice_1))
        # ax5_.yaxis.set_major_locator(MultipleLocator(3))
        # ax5_.set_yticklabels(['0','0','90','180','270'])
        # ax5_.set_ylabel('ground truth',fontsize=8)

        # ax6_ = plt.subplot(4,2,6)
        # # plt.contourf(label_calcium_squeeze)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        # plt.xticks(fontsize=7)
        # plt.yticks(fontsize=7)
        # h_label_calcium_ = plt.contourf(label_calcium_squeeze, cmap=plt.cm.coolwarm, levels=levels)
        # c_label_calcium_ = plt.colorbar(h_label_calcium_)
        # c_label_calcium_.ax.tick_params(labelsize=7)
        # ax6_.set_xlim((start_slice_1,end_slice_1))
        # ax6_.yaxis.set_major_locator(MultipleLocator(3))
        # ax6_.set_yticklabels(['0','0','90','180','270'])
        # # ax4_.set_ylabel('Ca label')
        # # print(1.0)

        # ax7_ = plt.subplot(4,2,7)
        # # plt.contourf(expert_lipid_squeeze)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        # plt.xticks(fontsize=7)
        # plt.yticks(fontsize=7)
        # h_expert_lipid_ = plt.contourf(expert_lipid_squeeze, cmap=plt.cm.coolwarm, levels=levels)
        # c_expert_lipid_ = plt.colorbar(h_expert_lipid_)
        # c_expert_lipid_.ax.tick_params(labelsize=7)
        # ax7_.set_xlim((start_slice_1,end_slice_1))
        # ax7_.yaxis.set_major_locator(MultipleLocator(3))
        # ax7_.set_yticklabels(['0','0','90','180','270'])
        # ax7_.set_ylabel('QAngio CT',fontsize=8)

        # ax8_ = plt.subplot(4,2,8)
        # # plt.contourf(expert_calcium_squeeze)
        # plt.contourf(shadow_1,cmap='Greens',levels=levels2)
        # plt.xticks(fontsize=7)
        # plt.yticks(fontsize=7)
        # h_expert_calcium_ = plt.contourf(expert_calcium_squeeze, cmap=plt.cm.coolwarm, levels=levels)
        # c_expert_calcium_ = plt.colorbar(h_expert_calcium_)
        # c_expert_calcium_.ax.tick_params(labelsize=7)
        # ax8_.set_xlim((start_slice_1,end_slice_1))
        # ax8_.yaxis.set_major_locator(MultipleLocator(3))
        # ax8_.set_yticklabels(['0','0','90','180','270'])
        # # ax5_.set_ylabel('expert')

        # # no need to further limit
        # # transversal_lipid = 0.7 # >=0.6
        # # transversal_calcium = 0.9 # >=0.8

        # fig_.suptitle(vessel)
        os.makedirs(os.path.join(chemo_png_save_path+'_'+str(angle_interval)+'deg'), exist_ok=True)
        plt.savefig(os.path.join(chemo_png_save_path+'_'+str(angle_interval)+'deg',vessel+'.png'))
    '''write image level based dice coefficient to excel'''
    # writer = pd.ExcelWriter(excel_p,engine='openpyxl')
    # book = openpyxl.load_workbook(writer.path)
    # writer.book = book
    # df_expert.to_excel(writer,'Expert',index=False)
    # df_pred.to_excel(writer,'3D2DUNet_BCE_WithoutStent',index=False)
    # writer.save()
    # writer.close()
    # plt.show()

def determin_lumen_center_(center,sample_line_idx,sample_freq,ct):
    for i,line in enumerate(sample_line_idx):
        angle = int(i*(1/sample_freq))
        for point in line:
            if angle<=45 or angle>=135:  # line_sample_idx: [row,col]
                if ct[point[0],point[1]] == 0:
                    continue
                rho = math.sqrt((point[0]-center[1])**2+(point[1]-center[0])**2)
                rho = int(np.floor(rho))
                if rho>=64:
                    center = lumen_center(ct,contour_idx=0) #[col,row]
                    return center

            if angle>45 and angle<135:  # line_sample_idx: [col,row]
                if ct[point[1],point[0]] == 0:
                    continue
                rho = math.sqrt((point[1]-center[1])**2+(point[0]-center[0])**2)
                rho = int(np.floor(rho))
                if rho>=64:
                    center = lumen_center(ct,contour_idx=0) #[col,row]
                    return center

    return center

def chemogram_with_lumen_center_ave(sample_freq,img_mode,idx_list,threshold,axis,pred_root_path,save_path,oriCT_root_path,sample_line_path,include_str=None,include_str_mask=None,root_path_mask=None,mask=None):
    sample_steps = int(180/sample_freq)
    # lumenCenter_sampleLine = np.load(r'\\vf-lkeb\lkeb$\Scratch\Xiaotong\lipid_plaque_detection\enlargedDLplaque\lumenCenter_sampleLine.npy',allow_pickle=True).item()
    lumenCenter_sampleLine = np.load(sample_line_path, allow_pickle=True).item()
    if os.path.exists(save_path):
        return

    for root, dirs, files in os.walk(pred_root_path):
        files = sorted(files)
        if len(files)!=0:
            print(root.split('/')[-1])
            if idx_list is None:
                last_slice_figure = int(re.findall(r"\d+",files[-1])[-1])
            else:
                last_slice_figure = idx_list[-1]
            sample_angles = 2*sample_steps
            chemogram = np.zeros((sample_angles,last_slice_figure+1))
        else:
            continue

        for name in files:
            if include_str is None:
                pass
            else:
                if include_str not in name:
                    continue
            # print(name)
            slice_idx = name.split('.')[0] # str
            slice_idx_figure = int(re.findall(r"\d+",slice_idx)[-1])
            if idx_list is not None:
                if slice_idx_figure not in idx_list:
                    continue
            # pred = imageio.imread(os.path.join(root,name))
            pred = nibabel.load(os.path.join(root,name)).get_fdata()
            pred = np.squeeze(pred)

            'co-registered label need it'
            # pred[pred==1] = 0
            # pred[pred==2] = 1

            '''calculate lumen center and sample line index'''
            ct_img = nibabel.load(os.path.join(oriCT_root_path,'CT-1plaque4mm'+str(int(re.findall(r"\d+",name)[-1]))+'.nii.gz')).get_fdata()

            ct_img = ct_img+1024
            if ct_img.sum()==0: # full-zeros CT slices without valid label
                continue
            ct_img[ct_img<0] = 0.1

            center = lumen_center(ct_img)
            if str(center) in lumenCenter_sampleLine.keys():
                pixel_index = lumenCenter_sampleLine[str(center)].copy()
            else:
                pixel_index = sample_line(center,sample_freq)
                center_update = determin_lumen_center_(center=center,sample_line_idx=pixel_index,sample_freq=sample_freq,ct=ct_img)
                if center_update!=center: # first lumen center calculation is wrong
                    center = center_update.copy() # update lumen center
                    pixel_index = sample_line(center,sample_freq) # update sample line index based on new lumen center
                tmp_dict = {str(center):pixel_index}
                lumenCenter_sampleLine.update(tmp_dict)
                # print(root.split('/')[-1]+': '+name)
            ''''''

            if include_str_mask is not None:
            # if mask is None: 
                mask = nibabel.load(os.path.join(root_path_mask,include_str_mask+str(slice_idx_figure).zfill(3)+'.nii.gz')).get_fdata()
                # mask_ori = nibabel.load(os.path.join(root_path_mask_ori,include_str_mask+str(slice_idx_figure).zfill(3)+'.nii.gz')).get_fdata() 
                # mask_ori_erode = np.zeros(mask_ori.shape)
                # for ite in range(mask_ori.shape[-1]):
                #     fill = scipy.ndimage.binary_fill_holes(mask_ori[:,:,ite])+0
                #     kernel = np.ones((30,30),np.uint8)
                #     mask_ori_erode[:,:,ite] = cv2.erode(fill.astype(np.float),kernel,1)
                #     mask_ori[:,:,ite] = mask_ori_erode[:,:,ite]*mask_ori[:,:,ite]
                pred = pred*mask
                # pred = pred*mask_ori
            else:
                # with fixed mask
                mask = np.ones((128,128))
                for i in range(128):
                    for j in range(128):
                        d = math.sqrt((i-center[1])*(i-center[1])+(j-center[0])*(j-center[0]))
                        r =int(np.round(d))
                        if r<=5:
                            mask[i,j] = 0
                mask_ct_img = (ct_img>0) +0
                mask = mask*mask_ct_img
                mask = np.expand_dims(mask,axis=-1)
                mask = np.repeat(mask,2,axis=-1)
                pred = pred*mask
            pred = pred[:,:,axis]

            pred[pred<0.1] = 0
            cnt = 0
            for angle in range(len(pixel_index)): #[0,180]
                cnt = cnt+1
                ''' part one '''
                if angle<=45 or angle>=135: # [row,col]
                    pred_pixel_values = pred[pixel_index[angle][:,0],pixel_index[angle][:,1]]
                    positive_flag = pred_pixel_values!=0
                    positive_flag = positive_flag+0

                    if positive_flag.sum()==0: 
                        chemogram[angle,slice_idx_figure] = -0.1
                        chemogram[180+angle,slice_idx_figure] = -0.1
                    else: 
                        first = []
                        second = []
                        third = []
                        forth = []
                        for idx in range(len(pixel_index[angle])):
                            if pixel_index[angle][idx][0]<=center[1]-1 and pixel_index[angle][idx][1]>=center[0]: # first section
                                first.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][0]<=center[1]-1 and pixel_index[angle][idx][1]<center[0]: # second section
                                second.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][0]>center[1]-1 and pixel_index[angle][idx][1]<center[0]: # third section
                                third.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][0]>center[1]-1 and pixel_index[angle][idx][1]>=center[0]: # forth section
                                forth.append(pixel_index[angle][idx])
                        first = np.array(first)
                        second = np.array(second)
                        third = np.array(third)
                        forth = np.array(forth)
                        if len(first)!=0 and len(third)!=0: # first-third
                            pred_pixel_values_first = pred[first[:,0],first[:,1]]
                            positive_flag_first = pred_pixel_values_first!=0
                            positive_flag_first = positive_flag_first+0
                            pred_pixel_values_first_ave = np.mean(pred_pixel_values_first[pred_pixel_values_first!=0])

                            pred_pixel_values_third = pred[third[:,0],third[:,1]]
                            positive_flag_third = pred_pixel_values_third!=0
                            positive_flag_third = positive_flag_third+0
                            pred_pixel_values_third_ave = np.mean(pred_pixel_values_third[pred_pixel_values_third!=0])


                            if img_mode=='binary':
                                if positive_flag_first.sum()==0: # no plaque on the angle
                                    chemogram[angle,slice_idx_figure] = -0.1
                                elif positive_flag_first.sum()!=0: # [0,1]
                                    chemogram[angle,slice_idx_figure] = 1

                                if positive_flag_third.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                elif positive_flag_third.sum()!=0:
                                    chemogram[180+angle,slice_idx_figure] = 1

                            elif img_mode=='consecutive':
                                if positive_flag_first.sum()==0: # no plaque on the angle
                                    chemogram[angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_first_ave>=0.8 and pred_pixel_values_first_ave<=1.0:
                                        chemogram[angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_first_ave>=0.6 and pred_pixel_values_first_ave<0.8:
                                        chemogram[angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_first_ave>=0.4 and pred_pixel_values_first_ave<0.6:
                                        chemogram[angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_first_ave>=0.2 and pred_pixel_values_first_ave<0.4:
                                        chemogram[angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_first_ave>=0.0 and pred_pixel_values_first_ave<0.2:
                                        chemogram[angle,slice_idx_figure] = 0.1
                            
                                if positive_flag_third.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_third_ave>=0.8 and pred_pixel_values_third_ave<=1.0:
                                        chemogram[180+angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_third_ave>=0.6 and pred_pixel_values_third_ave<0.8:
                                        chemogram[180+angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_third_ave>=0.4 and pred_pixel_values_third_ave<0.6:
                                        chemogram[180+angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_third_ave>=0.2 and pred_pixel_values_third_ave<0.4:
                                        chemogram[180+angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_third_ave>=0.0 and pred_pixel_values_third_ave<0.2:
                                        chemogram[180+angle,slice_idx_figure] = 0.1

                        elif len(second)!=0 and len(forth)!=0: # second-forth
                            pred_pixel_values_second = pred[second[:,0],second[:,1]]
                            positive_flag_second = pred_pixel_values_second!=0
                            positive_flag_second = positive_flag_second+0
                            pred_pixel_values_second_ave = np.mean(pred_pixel_values_second[pred_pixel_values_second!=0])

                            pred_pixel_values_forth = pred[forth[:,0],forth[:,1]]
                            positive_flag_forth = pred_pixel_values_forth!=0
                            positive_flag_forth = positive_flag_forth+0
                            pred_pixel_values_forth_ave = np.mean(pred_pixel_values_forth[pred_pixel_values_forth!=0])

                            if img_mode=='binary':
                                if positive_flag_second.sum()==0:
                                    chemogram[angle,slice_idx_figure] = -0.1
                                elif positive_flag_second.sum()!=0:
                                    chemogram[angle,slice_idx_figure] = 1

                                if positive_flag_forth.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                elif positive_flag_forth.sum()!=0:
                                    chemogram[180+angle,slice_idx_figure] = 1

                            elif img_mode=='consecutive':
                                if positive_flag_second.sum()==0:
                                    chemogram[angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_second_ave>=0.8 and pred_pixel_values_second_ave<=1.0:
                                        chemogram[angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_second_ave>=0.6 and pred_pixel_values_second_ave<0.8:
                                        chemogram[angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_second_ave>=0.4 and pred_pixel_values_second_ave<0.6:
                                        chemogram[angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_second_ave>=0.2 and pred_pixel_values_second_ave<0.4:
                                        chemogram[angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_second_ave>=0.0 and pred_pixel_values_second_ave<0.2:
                                        chemogram[angle,slice_idx_figure] = 0.1

                                if positive_flag_forth.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_forth_ave>=0.8 and pred_pixel_values_forth_ave<=1.0:
                                        chemogram[180+angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_forth_ave>=0.6 and pred_pixel_values_forth_ave<0.8:
                                        chemogram[180+angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_forth_ave>=0.4 and pred_pixel_values_forth_ave<0.6:
                                        chemogram[180+angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_forth_ave>=0.2 and pred_pixel_values_forth_ave<0.4:
                                        chemogram[180+angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_forth_ave>=0.0 and pred_pixel_values_forth_ave<0.2:
                                        chemogram[180+angle,slice_idx_figure] = 0.1


                ''' part two '''
                if angle>45 and angle<135: # [col,row]
                    pred_pixel_values = pred[pixel_index[angle][:,1],pixel_index[angle][:,0]]
                    positive_flag = pred_pixel_values!=0
                    positive_flag = positive_flag+0

                    if positive_flag.sum()==0:
                        chemogram[angle,slice_idx_figure] = -0.1
                        chemogram[180+angle,slice_idx_figure] = -0.1
                    else: 
                        first = []
                        second = []
                        third = []
                        forth = []
                        for idx in range(len(pixel_index[angle])):
                            if pixel_index[angle][idx][1]<=center[1] and pixel_index[angle][idx][0]>=center[0]: # first section
                                first.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][1]<=center[1] and pixel_index[angle][idx][0]<center[0]: # second section
                                second.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][1]>center[1] and pixel_index[angle][idx][0]<center[0]: # third section
                                third.append(pixel_index[angle][idx])
                            elif pixel_index[angle][idx][1]>center[1] and pixel_index[angle][idx][0]>=center[0]: # forth section
                                forth.append(pixel_index[angle][idx])
                        first = np.array(first)
                        second = np.array(second)
                        third = np.array(third)
                        forth = np.array(forth)
                        if len(first)!=0 and len(third)!=0: # first-third
                            pred_pixel_values_first = pred[first[:,1],first[:,0]]
                            positive_flag_first = pred_pixel_values_first!=0
                            positive_flag_first = positive_flag_first+0
                            pred_pixel_values_first_ave = np.mean(pred_pixel_values_first[pred_pixel_values_first!=0])

                            pred_pixel_values_third = pred[third[:,1],third[:,0]]
                            positive_flag_third = pred_pixel_values_third!=0
                            positive_flag_third = positive_flag_third+0
                            pred_pixel_values_third_ave = np.mean(pred_pixel_values_third[pred_pixel_values_third!=0])

                            if img_mode=='binary':
                                if positive_flag_first.sum()==0: # no plaque on the angle
                                    chemogram[angle,slice_idx_figure] = -0.1
                                elif positive_flag_first.sum()!=0: # [0,1]
                                    chemogram[angle,slice_idx_figure] = 1

                                if positive_flag_third.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                elif positive_flag_third.sum()!=0:
                                    chemogram[180+angle,slice_idx_figure] = 1

                            elif img_mode=='consecutive':
                                if positive_flag_first.sum()==0: # no plaque on the angle
                                    chemogram[angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_first_ave>=0.8 and pred_pixel_values_first_ave<=1.0:
                                        chemogram[angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_first_ave>=0.6 and pred_pixel_values_first_ave<0.8:
                                        chemogram[angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_first_ave>=0.4 and pred_pixel_values_first_ave<0.6:
                                        chemogram[angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_first_ave>=0.2 and pred_pixel_values_first_ave<0.4:
                                        chemogram[angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_first_ave>=0.0 and pred_pixel_values_first_ave<0.2:
                                        chemogram[angle,slice_idx_figure] = 0.1
                                
                                if positive_flag_third.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_third_ave>=0.8 and pred_pixel_values_third_ave<=1.0:
                                        chemogram[180+angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_third_ave>=0.6 and pred_pixel_values_third_ave<0.8:
                                        chemogram[180+angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_third_ave>=0.4 and pred_pixel_values_third_ave<0.6:
                                        chemogram[180+angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_third_ave>=0.2 and pred_pixel_values_third_ave<0.4:
                                        chemogram[180+angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_third_ave>=0.0 and pred_pixel_values_third_ave<0.2:
                                        chemogram[180+angle,slice_idx_figure] = 0.1

                        elif len(second)!=0 and len(forth)!=0: # second-forth
                            pred_pixel_values_second = pred[second[:,1],second[:,0]]
                            positive_flag_second = pred_pixel_values_second!=0
                            positive_flag_second = positive_flag_second+0
                            pred_pixel_values_second_ave = np.mean(pred_pixel_values_second[pred_pixel_values_second!=0])

                            pred_pixel_values_forth = pred[forth[:,1],forth[:,0]]
                            positive_flag_forth = pred_pixel_values_forth!=0
                            positive_flag_forth = positive_flag_forth+0
                            pred_pixel_values_forth_ave = np.mean(pred_pixel_values_forth[pred_pixel_values_forth!=0])

                            if img_mode=='binary':
                                if positive_flag_second.sum()==0:
                                    chemogram[angle,slice_idx_figure] = -0.1
                                elif positive_flag_second.sum()!=0:
                                    chemogram[angle,slice_idx_figure] = 1

                                if positive_flag_forth.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                elif positive_flag_forth.sum()!=0:
                                    chemogram[180+angle,slice_idx_figure] = 1

                            elif img_mode=='consecutive':
                                if positive_flag_second.sum()==0:
                                    chemogram[angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_second_ave>=0.8 and pred_pixel_values_second_ave<=1.0:
                                        chemogram[angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_second_ave>=0.6 and pred_pixel_values_second_ave<0.8:
                                        chemogram[angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_second_ave>=0.4 and pred_pixel_values_second_ave<0.6:
                                        chemogram[angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_second_ave>=0.2 and pred_pixel_values_second_ave<0.4:
                                        chemogram[angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_second_ave>=0.0 and pred_pixel_values_second_ave<0.2:
                                        chemogram[angle,slice_idx_figure] = 0.1

                                if positive_flag_forth.sum()==0:
                                    chemogram[180+angle,slice_idx_figure] = -0.1
                                else:
                                    if pred_pixel_values_forth_ave>=0.8 and pred_pixel_values_forth_ave<=1.0:
                                        chemogram[180+angle,slice_idx_figure] = 0.9  
                                    elif pred_pixel_values_forth_ave>=0.6 and pred_pixel_values_forth_ave<0.8:
                                        chemogram[180+angle,slice_idx_figure] = 0.7
                                    elif pred_pixel_values_forth_ave>=0.4 and pred_pixel_values_forth_ave<0.6:
                                        chemogram[180+angle,slice_idx_figure] = 0.5
                                    elif pred_pixel_values_forth_ave>=0.2 and pred_pixel_values_forth_ave<0.4:
                                        chemogram[180+angle,slice_idx_figure] = 0.3
                                    elif pred_pixel_values_forth_ave>=0.0 and pred_pixel_values_forth_ave<0.2:
                                        chemogram[180+angle,slice_idx_figure] = 0.1
    

        # imageio.imwrite(save_path,chemogram)
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        save_nii_gz(save_path,chemogram)
    # np.save(r'\\vf-lkeb\lkeb$\Scratch\Xiaotong\lipid_plaque_detection\enlargedDLplaque\lumenCenter_sampleLine.npy',lumenCenter_sampleLine)

def chemogram_generation_enlargedDL(save_path,pred_path,oriCT_root_path,sample_line_path): # with lumen center and old mask (lumen area)
    # save_path = r'Z:\lipid_plaque_detection\UNet_Keras\Experiment_expert_train_lumenCenter\chemogram\PN3115_UNet3D2D_ave_scapis'
    # pred_path = r'Z:\lipid_plaque_detection\UNet_Keras\Experiment_expert_train_lumenCenter\test_results_PN3115_unet3D2D_scapis'
    # oriCT_root_path = r'Z:\lipid_plaque_detection\scapis\deeplearning'

    test_vessels_name = os.listdir(pred_path)

    for vessel in test_vessels_name:
        ###lipid: axis=0  calcium: axis=1
        # if vessel!='02_LAD_final_DL' and vessel!='04_LAD_DLupdated':
        #     continue

        if os.path.exists(os.path.join(save_path,vessel+'_lipid.nii.gz')) or os.path.exists(os.path.join(save_path,vessel+'_calcium.nii.gz')):
            continue

        chemogram_with_lumen_center_ave(sample_freq=1,img_mode='consecutive',idx_list=None,threshold=[0,0.2,0.4,0.6,0.8,1.0],axis=0,#include_str_mask='Blockmask2CT',
                                        pred_root_path=os.path.join(pred_path,vessel),
                                        save_path=os.path.join(save_path,vessel+'_lipid.nii.gz'),
                                        oriCT_root_path=os.path.join(oriCT_root_path,vessel),
                                        sample_line_path=sample_line_path
                                        ),
                                        
                                        # root_path_mask=os.path.join(mask_path,vessel_mask))

        chemogram_with_lumen_center_ave(sample_freq=1,img_mode='consecutive',idx_list=None,threshold=[0,0.2,0.4,0.6,0.8,1.0],axis=1,#include_str_mask='Blockmask2CT',
                                        pred_root_path=os.path.join(pred_path,vessel),
                                        save_path=os.path.join(save_path,vessel+'_calcium.nii.gz'),
                                        oriCT_root_path=os.path.join(oriCT_root_path,vessel),
                                        sample_line_path=sample_line_path
                                        ),
                                        
                                        # root_path_mask=os.path.join(mask_path,vessel_mask))

def excel_template(excel_path, data_path):
    sheetname = 'scapis'

    data_col_1 = []
    data_col_2 = []

    for root, dirs, files in os.walk(data_path): 
        if root.split('/')[-1]!='02_LAD_final_DL' and root.split('/')[-1]!='04_LAD_DLupdated':
            continue
        data_col_1_vessel = []
        data_col_2_vessel = []
        for name in files:
            vessel_name = root.split('/')[-1]
            if 'CT-1plaque4mm' not in name:
                continue
            slice_idx = str(int(re.findall(r"\d+",name)[-1]))
            data_col_1_vessel.append(vessel_name)
            data_col_2_vessel.append(slice_idx)
        if len(data_col_1_vessel) !=0:
            data_col_2_vessel.sort(key=int)
            data_col_1.extend(data_col_1_vessel)
            data_col_2.extend(data_col_2_vessel)

    df = pd.DataFrame(data=list(zip(data_col_1,data_col_2)), columns=['VesselName','CTSliceIdx'])
    
    writer = pd.ExcelWriter(excel_path,engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    df.to_excel(writer,sheetname,index=False)
    writer.save()
    writer.close()

def BI(y_pred): # Burden index
    positive_points = np.sum(np.round(y_pred))
    burden_index = positive_points/12 # 360/30deg==12
    return burden_index

def excel_metrics_leision(pred,vessel,df,mode):
    leision_slice_idx = []
    row_idx = [] # corresponding to vessel
    cnt = 0
    for slice_idx in df['CTSliceIdx']:
        if np.isnan(slice_idx):
            continue
        if df['VesselName'][cnt]==vessel:
            leision_slice_idx.append(int(slice_idx))
            row_idx.append(cnt)
        cnt = cnt+1

    col_name = df.columns.tolist()
    col_idx = 3
    if 'LCBI_pred' not in col_name:
        col_name.insert(col_idx,'LCBI_pred')
        df = df.reindex(columns=col_name)
        col_idx = col_idx+1
    if 'CaBI_pred' not in col_name:
        col_name.insert(col_idx,'CaBI_pred')
        df = df.reindex(columns=col_name)
        col_idx = col_idx+1

    'angle-level evaluation'

    cnt_ = 0
    for col in range(pred.shape[1]):

        if col in leision_slice_idx: # a slice needed to be tested
            y_pred = pred[:,col].copy()
            bi_expert = BI(y_pred=y_pred) # burden index of predicted results

            if mode=='calcium':
                df['CaBI_pred'][row_idx[cnt_]] = np.round(bi_expert,2)

            elif mode=='lipid':
                df['LCBI_pred'][row_idx[cnt_]] = np.round(bi_expert,2)

            cnt_ = cnt_+1

    return df

def save_metrics(chemogram_pred_save_path,df,sheetname):
    test_vessels_name =  ['02_LAD_final_DL', '04_LAD_DLupdated']

    # for root, dirs, files in os.walk(chemogram_pred_save_path):
    #     for name in files:
    #         if name not in test_vessels_name:
    #             test_vessels_name.append(name.split('_')[0]+'_'+name.split('_')[1]+'_'+name.split('_')[2])

    test_vessels_name = sorted(test_vessels_name)

    vessel_idx = 0
    for vessel in test_vessels_name:
        print(vessel)

        pred_lipid = nibabel.load(os.path.join(chemogram_pred_save_path,vessel+'_lipid.nii.gz')).get_fdata()  # -0.1 or 1
        pred_calcium = nibabel.load(os.path.join(chemogram_pred_save_path,vessel+'_calcium.nii.gz')).get_fdata()

        angle_interval = 30

        pred_lipid[pred_lipid==-0.1] = 0
        pred_calcium[pred_calcium==-0.1] = 0

        'For DenseUNet'
        # show_thre_lipid = 0.6
        # show_thre_calcium = 0.6
        # pred_lipid[pred_lipid<show_thre_lipid+0.1] = 0
        # pred_calcium[pred_calcium<show_thre_calcium+0.1] = 0

        'For MaskRCNN (only the lipid boxes with probability >0.3 and the calcium boxes >0.5 are outputed)'

        ### squeezed chemogram
        pred_lipid_squeeze = np.zeros((int(pred_lipid.shape[0]/angle_interval),pred_lipid.shape[1]))   
        pred_calcium_squeeze = np.zeros((int(pred_calcium.shape[0]/angle_interval),pred_calcium.shape[1]))  
        # row_cnt = 0
        for row in range(pred_lipid.shape[0]):
            # row_cnt = row_cnt+1
            tmp_pred_lipid = pred_lipid.copy()
            tmp_pred_calcium = pred_calcium.copy()

            if (row+1)%angle_interval==0:
                pred_lipid_squeeze[int(row/angle_interval),:] = np.mean(tmp_pred_lipid[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)
                pred_calcium_squeeze[int(row/angle_interval),:] = np.mean(tmp_pred_calcium[int(row/angle_interval)*angle_interval:int((row/angle_interval+1))*angle_interval,:],axis=0)

        pred_lipid_squeeze = np.round(pred_lipid_squeeze,decimals=1)
        pred_calcium_squeeze = np.round(pred_calcium_squeeze,decimals=1)

        df = excel_metrics_leision(pred=pred_lipid_squeeze,vessel=vessel,df=df,mode='lipid')
        df = excel_metrics_leision(pred=pred_calcium_squeeze,vessel=vessel,df=df,mode='calcium')
        vessel_idx = vessel_idx+1

    writer = pd.ExcelWriter(r'\\vf-lkeb\lkeb$\Scratch\Xiaotong\lipid_plaque_detection\scapis_LCBI_CaBI_2cases.xlsx',engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    df.to_excel(writer,sheetname,index=False)
    writer.save()
    writer.close()

if __name__ == '__main__':

    # chemogram_plot_RGB_expert_IVUS(
    #                                chemogram_save_path=r'Z:\lipid_plaque_detection\MaskRCNN_Pytorch\outputs\chemogram_test\baseline2_testForScapis'
    #                                )
    # chemogram_generation_enlargedDL()
    # chemogram_plot_final(
    #                      chemogram_save_path=[r'Z:\lipid_plaque_detection\UNet_Keras\Experiment_expert_train_lumenCenter\chemogram\PN3115_UNet3D2D_ave_scapis', 
    #                                           r'Z:\lipid_plaque_detection\MaskRCNN_Pytorch\outputs\chemogram_test\baseline2_testForScapis'],
    #                      chemo_png_save_path=r'Z:\lipid_plaque_detection\scapis_chemo_comparison\unet_rcnn_png')   

    excel_path = r'Z:\lipid_plaque_detection\scapis_LCBI_CaBI_2cases.xlsx'
    # excel_template(excel_path=excel_path,
    #                 data_path=r'Z:\lipid_plaque_detection\scapis\deeplearning')
    df = pd.read_excel(excel_path,sheet_name='scapis',engine='openpyxl')
    # save_metrics(chemogram_pred_save_path=r'Z:\lipid_plaque_detection\UNet_Keras\Experiment_expert_train_lumenCenter\chemogram\PN3115_UNet3D2D_ave_scapis', df=df, sheetname='metrics_DenseUNet')
    save_metrics(chemogram_pred_save_path=r'Z:\lipid_plaque_detection\MaskRCNN_Pytorch\outputs\chemogram_test\baseline2_testForScapis', df=df, sheetname='metrics_MaskRCNN')